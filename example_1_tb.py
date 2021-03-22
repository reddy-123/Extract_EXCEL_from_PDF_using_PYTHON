import openpyxl

# Import the excel file
wb = openpyxl.load_workbook(r'excel\pbc_tb.xlsx', data_only=True)
sheet = wb.get_sheet_by_name('tb')
last_row = sheet.max_row
import_tb = [['gl', 'description', 'net']]

for row in range(4, last_row + 1):
    row = str(row)   
    gl_description = sheet['B'+row].value
    if gl_description == 'Total':
        continue

  

if rounding != 1:
    import_tb.append(['9999', 'rounding', -rounding])

# Create a new tab with the columns “acct”, “description”, and “balance”
import_sheet = wb.create_sheet(title=r'tb_import', index=0)

for num, (gl, desc, net) in enumerate(import_tb):
    import_sheet['A' + str(num + 1)].value = gl
    import_sheet['B' + str(num + 1)].value = desc
    import_sheet['C' + str(num + 1)].value = net

# Save to a new file
wb.save(r'excel\tb_import.xlsx')
